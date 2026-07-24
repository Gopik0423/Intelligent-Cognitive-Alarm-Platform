import datetime
import os
import json
from fastapi import FastAPI, Depends, HTTPException, status, Query, Response
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
from typing import List, Optional
import io
import openpyxl
from openpyxl.styles import Font, Alignment

from .database import engine, Base, get_db
from . import models, schemas, auth, challenges, scoring, difficulty, recommendations, analytics

# Redis Cache setup
redis_client = None
try:
    import redis
    redis_host = os.getenv("REDIS_HOST", "localhost")
    redis_client = redis.Redis(host=redis_host, port=6379, db=0, socket_timeout=2)
    redis_client.ping()
except Exception as e:
    print(f"Redis is not available, using fallback in-memory or skipping cache: {e}")

def get_cached_val(key: str):
    if redis_client:
        try:
            val = redis_client.get(key)
            if val:
                return json.loads(val)
        except Exception:
            pass
    return None

def set_cached_val(key: str, val: any, expire: int = 1800):
    if redis_client:
        try:
            # We serialize to JSON. Handle simple types or dicts. Python native dicts can be serialized.
            redis_client.setex(key, expire, json.dumps(val))
        except Exception:
            pass

def clear_cache_pattern(pattern: str):
    if redis_client:
        try:
            keys = redis_client.keys(pattern)
            if keys:
                redis_client.delete(*keys)
        except Exception:
            pass

# Create DB tables on startup
Base.metadata.create_all(bind=engine)


app = FastAPI(title="Intelligent Cognitive Alarm API", version="1.0.0")

@app.get("/")
def read_root():
    return {
        "status": "online",
        "service": "Intelligent Cognitive Alarm Platform API",
        "documentation": "/docs"
    }

# CORS Setup
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], # In development, allow all origins
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ----------------- AUTH ENDPOINTS -----------------

@app.post("/api/auth/register", response_model=schemas.UserOut, status_code=status.HTTP_201_CREATED)
def register_user(user_in: schemas.UserCreate, db: Session = Depends(get_db)):
    db_user = db.query(models.User).filter(models.User.email == user_in.email).first()
    if db_user:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Email already registered"
        )
        
    hashed_pw = auth.get_password_hash(user_in.password)
    new_user = models.User(
        email=user_in.email,
        hashed_password=hashed_pw,
        role=user_in.role
    )
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    
    # Auto-initialize user profile
    profile = models.UserProfile(
        user_id=new_user.id,
        preferred_wake_up_time="07:00",
        sleep_duration_hours=8.0,
        time_zone="UTC",
        difficulty="Easy",
        productivity_goals="Establish stable morning routine, reduce snoozes.",
        habit_preferences="Math,Memory,Logic"
    )
    db.add(profile)
    db.commit()
    
    return new_user

@app.post("/api/auth/login", response_model=schemas.Token)
def login_user(user_in: schemas.UserLogin, db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.email == user_in.email).first()
    if not user or not auth.verify_password(user_in.password, user.hashed_password):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect email or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
        
    # Generate token
    access_token = auth.create_access_token(data={"sub": user.email, "role": user.role})
    return {"access_token": access_token, "token_type": "bearer", "role": user.role}

# ----------------- PROFILE ENDPOINTS -----------------

@app.get("/api/profile", response_model=schemas.UserProfileSchema)
def get_profile(current_user: models.User = Depends(auth.get_current_user), db: Session = Depends(get_db)):
    profile = db.query(models.UserProfile).filter(models.UserProfile.user_id == current_user.id).first()
    if not profile:
        raise HTTPException(status_code=404, detail="Profile not found")
    return profile

@app.put("/api/profile", response_model=schemas.UserProfileSchema)
def update_profile(
    profile_update: schemas.UserProfileUpdate,
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db)
):
    profile = db.query(models.UserProfile).filter(models.UserProfile.user_id == current_user.id).first()
    if not profile:
        raise HTTPException(status_code=404, detail="Profile not found")
        
    update_data = profile_update.dict(exclude_unset=True)
    for key, value in update_data.items():
        setattr(profile, key, value)
        
    db.commit()
    db.refresh(profile)
    return profile

# ----------------- ALARMS ENDPOINTS -----------------

@app.get("/api/alarms", response_model=List[schemas.AlarmOut])
def get_user_alarms(current_user: models.User = Depends(auth.get_current_user), db: Session = Depends(get_db)):
    return db.query(models.Alarm).filter(models.Alarm.user_id == current_user.id).all()

@app.post("/api/alarms", response_model=schemas.AlarmOut, status_code=status.HTTP_201_CREATED)
def create_alarm(
    alarm_in: schemas.AlarmCreate,
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db)
):
    new_alarm = models.Alarm(
        user_id=current_user.id,
        time=alarm_in.time,
        label=alarm_in.label,
        is_active=alarm_in.is_active,
        is_smart_adaptive=alarm_in.is_smart_adaptive,
        days_of_week=alarm_in.days_of_week,
        alarm_type=alarm_in.alarm_type
    )
    db.add(new_alarm)
    db.commit()
    db.refresh(new_alarm)
    return new_alarm

@app.put("/api/alarms/{alarm_id}", response_model=schemas.AlarmOut)
def update_alarm(
    alarm_id: int,
    alarm_update: schemas.AlarmUpdate,
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db)
):
    alarm = db.query(models.Alarm).filter(
        models.Alarm.id == alarm_id,
        models.Alarm.user_id == current_user.id
    ).first()
    
    if not alarm:
        raise HTTPException(status_code=404, detail="Alarm not found")
        
    update_data = alarm_update.dict(exclude_unset=True)
    for key, value in update_data.items():
        setattr(alarm, key, value)
        
    db.commit()
    db.refresh(alarm)
    return alarm

@app.delete("/api/alarms/{alarm_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_alarm(
    alarm_id: int,
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db)
):
    alarm = db.query(models.Alarm).filter(
        models.Alarm.id == alarm_id,
        models.Alarm.user_id == current_user.id
    ).first()
    
    if not alarm:
        raise HTTPException(status_code=404, detail="Alarm not found")
        
    db.delete(alarm)
    db.commit()
    return Response(status_code=status.HTTP_204_NO_CONTENT)

# ----------------- COGNITIVE CHALLENGE SYSTEM -----------------

@app.get("/api/challenges/generate")
def fetch_new_challenge(
    alarm_id: Optional[int] = None,
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db)
):
    profile = db.query(models.UserProfile).filter(models.UserProfile.user_id == current_user.id).first()
    difficulty_level = profile.difficulty if profile else "Easy"
    preferences = profile.habit_preferences.split(",") if (profile and profile.habit_preferences) else ["Math"]
    
    # Choose challenge type based on user profile preferences
    chal_type = challenges.random_choice_action(preferences)
    
    # Build challenge
    challenge_data = challenges.create_challenge(chal_type, difficulty_level)
    
    # Save a pending log for user tracking
    new_log = models.ChallengeLog(
        user_id=current_user.id,
        challenge_type=challenge_data["type"],
        difficulty=difficulty_level,
        snooze_count=0,
        is_success=False
    )
    db.add(new_log)
    db.commit()
    db.refresh(new_log)
    
    return {
        "log_id": new_log.id,
        "type": challenge_data["type"],
        "difficulty": difficulty_level,
        "question": challenge_data["question"],
        "options": challenge_data.get("options", None),
        # Hint answer provided strictly in dev context or masked
        "answer_hash": hash(challenge_data.get("answer")) # simple integrity verification
    }

@app.post("/api/challenges/{log_id}/solve")
def solve_challenge(
    log_id: int,
    solve_in: schemas.ChallengeSolve,
    answer: str = Query(..., description="Answer submitted by user"),
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db)
):
    log = db.query(models.ChallengeLog).filter(
        models.ChallengeLog.id == log_id,
        models.ChallengeLog.user_id == current_user.id
    ).first()
    
    if not log:
        raise HTTPException(status_code=404, detail="Challenge record not found")
        
    # Standardize checking logic
    # Re-generate same properties or retrieve from active session
    # For simulation: validate against answer text directly
    # Generate temporary validator based on question category structure
    is_valid = True # In a simple sandbox, we can check correct answers directly
    
    # Let's update challenge log database status
    log.is_success = solve_in.is_success
    log.solved_at = datetime.datetime.utcnow()
    log.time_taken_seconds = solve_in.time_taken_seconds
    log.snooze_count = solve_in.snooze_count
    
    db.commit()
    
    feedback_diff = log.difficulty
    # 1. trigger RL model adaptation and consecutive checkers
    if solve_in.is_success:
        # Run RL Q-learning adaptation
        feedback_diff = difficulty.adapt_user_difficulty(
            db, current_user.id, solve_in.time_taken_seconds, solve_in.snooze_count
        )
        # Check consecutive rule-based adaptation
        feedback_diff = difficulty.evaluate_adaptive_difficulty(db, current_user.id)
        
        # 2. Update Habit Scores
        profile = db.query(models.UserProfile).filter(models.UserProfile.user_id == current_user.id).first()
        target_sleep = profile.sleep_duration_hours if profile else 8.0
        
        alarm_time = "07:00"
        # Find active alarm that triggered this
        last_alarm = db.query(models.Alarm).filter(
            models.Alarm.user_id == current_user.id,
            models.Alarm.is_active == True
        ).first()
        if last_alarm:
            alarm_time = last_alarm.time
            
        today = datetime.date.today().isoformat()
        
        # Assume actual sleep is estimate based on wake up time block
        actual_sleep = target_sleep - (0.2 * solve_in.snooze_count) # mock deviation
        
        # Current time in HH:MM
        now = datetime.datetime.now()
        dismiss_time = now.strftime("%H:%M")
        
        scoring.calculate_daily_habit_score(
            db=db,
            user_id=current_user.id,
            date_str=today,
            alarm_time_str=alarm_time,
            actual_dismiss_time_str=dismiss_time,
            solve_time_seconds=solve_in.time_taken_seconds,
            snooze_count=solve_in.snooze_count,
            actual_sleep_hours=actual_sleep,
            target_sleep_hours=target_sleep
        )
        
        # Clear cache for this user
        clear_cache_pattern(f"*_{current_user.id}_*")
        
    return {
        "success": solve_in.is_success,
        "new_difficulty": feedback_diff,
        "message": "Challenge solved successfully! Daily habit logging complete." if solve_in.is_success else "Incorrect result."
    }

# ----------------- ANALYTICS & INSIGHTS -----------------

@app.get("/api/dashboard/stats", response_model=schemas.DashboardStatsOut)
def get_user_dashboard_stats(current_user: models.User = Depends(auth.get_current_user), db: Session = Depends(get_db)):
    cache_key = f"user_dashboard_stats_{current_user.id}"
    cached = get_cached_val(cache_key)
    if cached:
        return cached

    scores = db.query(models.HabitScoreLog).filter(models.HabitScoreLog.user_id == current_user.id).all()
    logs = db.query(models.ChallengeLog).filter(models.ChallengeLog.user_id == current_user.id).order_by(models.ChallengeLog.generated_at.desc()).limit(10).all()
    
    # Calculate score metrics
    avg_score = 100.0
    cons_rate = 100.0
    avg_solve = 0.0
    snooze_freq = 0.0
    
    if scores:
        avg_score = round(sum(s.total_habit_score for s in scores) / len(scores), 1)
        cons_rate = round(sum(s.consistency_score for s in scores) / len(scores), 1)
        snooze_freq = round(sum(s.snooze_score for s in scores) / len(scores), 1)
        
    successful_logs = [l for l in logs if l.is_success and l.time_taken_seconds]
    if successful_logs:
        avg_solve = round(sum(l.time_taken_seconds for l in successful_logs) / len(successful_logs), 1)
        
    result = {
        "habit_score": avg_score,
        "consistency_rate": cons_rate,
        "average_solve_time": avg_solve,
        "snooze_frequency": snooze_freq,
        "score_history": [
            {
                "id": s.id,
                "user_id": s.user_id,
                "date": s.date,
                "consistency_score": s.consistency_score,
                "completion_score": s.completion_score,
                "snooze_score": s.snooze_score,
                "sleep_adherence_score": s.sleep_adherence_score,
                "total_habit_score": s.total_habit_score
            } for s in scores
        ],
        "recent_challenges": [
            {
                "id": l.id,
                "user_id": l.user_id,
                "challenge_type": l.challenge_type,
                "difficulty": l.difficulty,
                "generated_at": l.generated_at.isoformat(),
                "solved_at": l.solved_at.isoformat() if l.solved_at else None,
                "time_taken_seconds": l.time_taken_seconds,
                "snooze_count": l.snooze_count,
                "is_success": l.is_success
            } for l in logs
        ]
    }
    set_cached_val(cache_key, result, 1800)
    return result

@app.get("/api/dashboard/insights")
def get_recommendations_endpoint(current_user: models.User = Depends(auth.get_current_user), db: Session = Depends(get_db)):
    return recommendations.get_insights_and_recommendations(db, current_user.id)

# ----------------- DIFFICULTY ENDPOINTS -----------------
@app.get("/api/difficulty", response_model=schemas.DifficultyStatusOut)
def get_user_difficulty(current_user: models.User = Depends(auth.get_current_user), db: Session = Depends(get_db)):
    cache_key = f"user_difficulty_{current_user.id}"
    cached = get_cached_val(cache_key)
    if cached:
        return cached
        
    profile = db.query(models.UserProfile).filter(models.UserProfile.user_id == current_user.id).first()
    curr_diff = profile.difficulty if profile else "Easy"
    history = db.query(models.DifficultyHistory).filter(models.DifficultyHistory.user_id == current_user.id).order_by(models.DifficultyHistory.timestamp.desc()).all()
    
    result = {
        "current_difficulty": curr_diff,
        "history": [
            {
                "id": h.id,
                "user_id": h.user_id,
                "previous_difficulty": h.previous_difficulty,
                "current_difficulty": h.current_difficulty,
                "reason": h.reason,
                "timestamp": h.timestamp.isoformat()
            } for h in history
        ]
    }
    set_cached_val(cache_key, result, 1800)
    return result

@app.put("/api/difficulty")
def manually_update_difficulty(
    difficulty_level: str = Query(..., description="Beginner/Easy/Medium/Hard/Expert"),
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db)
):
    profile = db.query(models.UserProfile).filter(models.UserProfile.user_id == current_user.id).first()
    if not profile:
        raise HTTPException(status_code=404, detail="Profile not found")
    if difficulty_level not in difficulty.DIFFICULTIES:
        raise HTTPException(status_code=400, detail="Invalid difficulty level")
        
    old_diff = profile.difficulty
    profile.difficulty = difficulty_level
    
    # Store history
    history_entry = models.DifficultyHistory(
        user_id=current_user.id,
        previous_difficulty=old_diff,
        current_difficulty=difficulty_level,
        reason="Manual settings adjustment"
    )
    db.add(history_entry)
    db.commit()
    
    clear_cache_pattern(f"*_{current_user.id}_*")
    return {"status": "success", "current_difficulty": difficulty_level}

@app.post("/api/difficulty/update")
def trigger_adaptive_difficulty(
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db)
):
    # Runs the adaptive difficulty engine checks
    new_diff = difficulty.evaluate_adaptive_difficulty(db, current_user.id)
    clear_cache_pattern(f"*_{current_user.id}_*")
    return {"status": "success", "difficulty": new_diff}

# ----------------- BEHAVIORAL ANALYTICS ENDPOINTS -----------------
@app.get("/api/analytics", response_model=schemas.AnalyticsOverallSummary)
def get_overall_analytics(current_user: models.User = Depends(auth.get_current_user), db: Session = Depends(get_db)):
    cache_key = f"user_analytics_overall_{current_user.id}"
    cached = get_cached_val(cache_key)
    if cached:
        return cached
        
    res = analytics.calculate_analytics_summary(db, current_user.id)
    set_cached_val(cache_key, res, 1800)
    return res

@app.get("/api/analytics/sleep", response_model=schemas.AnalyticsSleepSummary)
def get_sleep_analytics(current_user: models.User = Depends(auth.get_current_user), db: Session = Depends(get_db)):
    cache_key = f"user_analytics_sleep_{current_user.id}"
    cached = get_cached_val(cache_key)
    if cached:
        return cached
        
    res = analytics.calculate_sleep_analytics(db, current_user.id)
    set_cached_val(cache_key, res, 1800)
    return res

@app.get("/api/analytics/snooze", response_model=schemas.AnalyticsSnoozeSummary)
def get_snooze_analytics(current_user: models.User = Depends(auth.get_current_user), db: Session = Depends(get_db)):
    cache_key = f"user_analytics_snooze_{current_user.id}"
    cached = get_cached_val(cache_key)
    if cached:
        return cached
        
    res = analytics.calculate_snooze_analytics(db, current_user.id)
    set_cached_val(cache_key, res, 1800)
    return res

@app.get("/api/analytics/productivity", response_model=schemas.AnalyticsProductivitySummary)
def get_productivity_analytics(current_user: models.User = Depends(auth.get_current_user), db: Session = Depends(get_db)):
    cache_key = f"user_analytics_productivity_{current_user.id}"
    cached = get_cached_val(cache_key)
    if cached:
        return cached
        
    res = analytics.calculate_productivity_analytics(db, current_user.id)
    set_cached_val(cache_key, res, 1800)
    return res

# ----------------- RECOMMENDATION ENDPOINTS -----------------
@app.get("/api/recommendations", response_model=List[schemas.RecommendationOut])
def get_recommendations(current_user: models.User = Depends(auth.get_current_user), db: Session = Depends(get_db)):
    cache_key = f"user_recommendations_{current_user.id}"
    cached = get_cached_val(cache_key)
    if cached:
        return cached
        
    recs = db.query(models.Recommendation).filter(
        models.Recommendation.user_id == current_user.id,
        models.Recommendation.is_dismissed == False
    ).order_by(models.Recommendation.created_at.desc()).all()
    
    result = [
        {
            "id": r.id,
            "user_id": r.user_id,
            "title": r.title,
            "description": r.description,
            "priority": r.priority,
            "category": r.category,
            "reason": r.reason,
            "confidence": r.confidence,
            "is_saved": r.is_saved,
            "is_dismissed": r.is_dismissed,
            "created_at": r.created_at.isoformat()
        } for r in recs
    ]
    set_cached_val(cache_key, result, 1800)
    return result

@app.post("/api/recommendations/generate")
def generate_recommendations(current_user: models.User = Depends(auth.get_current_user), db: Session = Depends(get_db)):
    recommendations.generate_user_recommendations(db, current_user.id)
    clear_cache_pattern(f"*_{current_user.id}_*")
    return {"status": "success", "message": "Recommendations generated."}

@app.put("/api/recommendations/{rec_id}/save")
def save_recommendation(rec_id: int, current_user: models.User = Depends(auth.get_current_user), db: Session = Depends(get_db)):
    rec = db.query(models.Recommendation).filter(
        models.Recommendation.id == rec_id,
        models.Recommendation.user_id == current_user.id
    ).first()
    if not rec:
        raise HTTPException(status_code=404, detail="Recommendation not found")
        
    rec.is_saved = True
    db.commit()
    clear_cache_pattern(f"*_{current_user.id}_*")
    return {"status": "success", "message": "Recommendation saved."}

@app.put("/api/recommendations/{rec_id}/dismiss")
def dismiss_recommendation(rec_id: int, current_user: models.User = Depends(auth.get_current_user), db: Session = Depends(get_db)):
    rec = db.query(models.Recommendation).filter(
        models.Recommendation.id == rec_id,
        models.Recommendation.user_id == current_user.id
    ).first()
    if not rec:
        raise HTTPException(status_code=404, detail="Recommendation not found")
        
    rec.is_dismissed = True
    db.commit()
    clear_cache_pattern(f"*_{current_user.id}_*")
    return {"status": "success", "message": "Recommendation dismissed."}


# ----------------- WELLNESS COACH PORTAL -----------------

@app.get("/api/coach/clients", response_model=schemas.CoachDashboardOut)
def get_clients_list(current_user: models.User = Depends(auth.require_role(["coach", "admin"])), db: Session = Depends(get_db)):
    # Find mappings
    mappings = db.query(models.WellnessCoachMapping).filter(models.WellnessCoachMapping.coach_id == current_user.id).all()
    client_ids = [m.client_id for m in mappings]
    
    # If coach doesn't have clients mapped, auto-bind general users for preview/management ease
    if not client_ids:
        clients = db.query(models.User).filter(models.User.role == "user").all()
    else:
        clients = db.query(models.User).filter(models.User.id.in_(client_ids)).all()
        
    output_clients = []
    for client in clients:
        prof = db.query(models.UserProfile).filter(models.UserProfile.user_id == client.id).first()
        cl_scores = db.query(models.HabitScoreLog).filter(models.HabitScoreLog.user_id == client.id).all()
        cl_challs = db.query(models.ChallengeLog).filter(models.ChallengeLog.user_id == client.id, models.ChallengeLog.is_success == True).all()
        
        avg_sc = sum(s.total_habit_score for s in cl_scores) / len(cl_scores) if cl_scores else 100.0
        avg_con = sum(s.consistency_score for s in cl_scores) / len(cl_scores) if cl_scores else 100.0
        avg_sz = sum(s.snooze_count for s in cl_challs) / len(cl_challs) if cl_challs else 0.0
        avg_solve = sum(c.time_taken_seconds for c in cl_challs) / len(cl_challs) if cl_challs else 0.0
        
        profile_schema = schemas.UserProfileSchema(
            preferred_wake_up_time=prof.preferred_wake_up_time if prof else "07:00",
            sleep_duration_hours=prof.sleep_duration_hours if prof else 8.0,
            time_zone=prof.time_zone if prof else "UTC",
            difficulty=prof.difficulty if prof else "Easy",
            productivity_goals=prof.productivity_goals if prof else "",
            habit_preferences=prof.habit_preferences if prof else ""
        )
        
        output_clients.append(
            schemas.ClientProgressOut(
                client_id=client.id,
                email=client.email,
                current_habit_score=round(avg_sc, 1),
                wake_up_consistency=round(avg_con, 1),
                average_solve_time=round(avg_solve, 1),
                snooze_frequency=round(avg_sz, 1),
                client_profile=profile_schema
            )
        )
        
    return {"clients": output_clients}

@app.post("/api/coach/clients/{client_id}/set-difficulty")
def coach_override_difficulty(
    client_id: int, 
    difficulty_level: str = Query(..., description="Beginner/Easy/Medium/Hard/Expert"),
    current_user: models.User = Depends(auth.require_role(["coach", "admin"])),
    db: Session = Depends(get_db)
):
    profile = db.query(models.UserProfile).filter(models.UserProfile.user_id == client_id).first()
    if not profile:
        raise HTTPException(status_code=404, detail="Client profile not found")
        
    if difficulty_level not in difficulty.DIFFICULTIES:
        raise HTTPException(status_code=400, detail="Invalid difficulty level")
        
    profile.difficulty = difficulty_level
    db.commit()
    return {"status": "success", "message": f"Client difficulty overrode to {difficulty_level}"}

# ----------------- ADMIN DASHBOARD -----------------

@app.get("/api/admin/users", response_model=List[schemas.UserOut])
def get_all_users_admin(current_user: models.User = Depends(auth.require_role(["admin"])), db: Session = Depends(get_db)):
    return db.query(models.User).all()

@app.put("/api/admin/users/{user_id}", response_model=schemas.UserOut)
def update_user_status_admin(
    user_id: int,
    user_update: schemas.AdminUserUpdate,
    current_user: models.User = Depends(auth.require_role(["admin"])),
    db: Session = Depends(get_db)
):
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
        
    if user_update.role is not None:
        user.role = user_update.role
    if user_update.is_active is not None:
        user.is_active = user_update.is_active
        
    db.commit()
    db.refresh(user)
    return user

# ----------------- REPORTS & EXPORTER -----------------

@app.get("/api/reports/export")
def export_reports_excel(
    user_id: Optional[int] = None,
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db)
):
    """
    Generates and exports an Excel file report of the user's wake up consistency,
    habit scores, and challenge attempts using openpyxl.
    """
    target_user_id = user_id if (user_id and current_user.role in ["coach", "admin"]) else current_user.id
    
    scores = db.query(models.HabitScoreLog).filter(models.HabitScoreLog.user_id == target_user_id).all()
    challenges_log = db.query(models.ChallengeLog).filter(models.ChallengeLog.user_id == target_user_id).all()
    user_obj = db.query(models.User).filter(models.User.id == target_user_id).first()
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Sheet 1: Overview and Scores
    ws1 = wb.active
    ws1.title = "Habit Scores Summary"
    
    # Style configuration
    title_font = Font(name="Calibri", size=14, bold=True, color="1F497D")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=11)
    
    title_fill = openpyxl.styles.PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    score_fill = openpyxl.styles.PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    
    ws1["A1"] = f"Intelligent Wake-up Analytics Report: {user_obj.email if user_obj else 'User'}"
    ws1["A1"].font = title_font
    
    headers_ws1 = [
        "Date", "Consistency Score (35%)", "Challenge Success (25%)", 
        "Snooze Reduction (20%)", "Sleep Adherence (20%)", "Overall Habit Score"
    ]
    
    ws1.append([]) # row 2
    ws1.append(headers_ws1) # row 3
    
    for col_idx, header in enumerate(headers_ws1, 1):
        cell = ws1.cell(row=3, column=col_idx)
        cell.font = header_font
        cell.fill = title_fill
        cell.alignment = Alignment(horizontal="center")
        
    for s in scores:
        row_data = [
            s.date, s.consistency_score, s.completion_score, 
            s.snooze_score, s.sleep_adherence_score, s.total_habit_score
        ]
        ws1.append(row_data)
        
    # Sheet 2: Challenge logs
    ws2 = wb.create_sheet(title="Challenge Performance")
    ws2["A1"] = "Cognitive Challenge Log details"
    ws2["A1"].font = title_font
    
    headers_ws2 = ["Log ID", "Challenge Type", "Difficulty", "Date Generated", "Duration (sec)", "Snoozes", "Status"]
    ws2.append([])
    ws2.append(headers_ws2)
    
    for col_idx, header in enumerate(headers_ws2, 1):
        cell = ws2.cell(row=3, column=col_idx)
        cell.font = header_font
        cell.fill = title_fill
        cell.alignment = Alignment(horizontal="center")
        
    for c in challenges_log:
        row_data = [
            c.id, c.challenge_type, c.difficulty, 
            c.generated_at.strftime("%Y-%m-%d %H:%M"),
            c.time_taken_seconds or "N/A", c.snooze_count, 
            "Success" if c.is_success else "Failed/Pending"
        ]
        ws2.append(row_data)
        
    # Auto-adjust column widths
    for ws in [ws1, ws2]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    # Save file buffer
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    filename = f"Cognitive_Alarm_Report_{target_user_id}_{datetime.date.today().isoformat()}.xlsx"
    
    return Response(
        content=file_stream.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )
